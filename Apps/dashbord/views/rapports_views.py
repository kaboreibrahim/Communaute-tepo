from decimal import Decimal
from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from django.http import HttpResponse
from django.db.models import Count, Q, Avg, Sum
from django.utils import timezone
from datetime import datetime, date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from Apps.cotisations.models import (
    PERSON_TRACKING_STATUS_LABELS,
    Cotisation,
    CotisationPersonne,
    Paiement,
    compute_remaining_amount,
    resolve_person_tracking_status,
)
from Apps.person.models import Person
from Apps.families.models import Family
from Apps.villages.models import Village, Infrastructure
from Apps.dashbord.security import (
    filter_person_queryset_for_user,
    is_limited_data_entry_agent,
)


def _visible_persons(request):
    return filter_person_queryset_for_user(
        Person.objects.filter(deleted__isnull=True),
        request.user,
    )


def _visible_families(request):
    qs = Family.objects.filter(deleted__isnull=True)
    if is_limited_data_entry_agent(request.user):
        qs = qs.filter(
            membres__deleted__isnull=True,
            membres__created_by=request.user,
        ).distinct()
    return qs


def _visible_villages(request):
    qs = Village.objects.filter(deleted__isnull=True)
    if is_limited_data_entry_agent(request.user):
        qs = qs.filter(
            familles__membres__deleted__isnull=True,
            familles__membres__created_by=request.user,
        ).distinct()
    return qs


@login_required
def rapports_dashboard(request):
    """
    Vue principale pour la page des rapports
    """
    context = {
        'title': 'Rapports et Exportations',
        'stats': {
            'total_personnes': _visible_persons(request).count(),
            'total_familles': _visible_families(request).count(),
            'total_villages': _visible_villages(request).count(),
            'total_infrastructures': Infrastructure.objects.filter(deleted__isnull=True).count(),
            'total_cotisations': Cotisation.objects.count(),
            'paiements_cotisation_pending': Paiement.objects.filter(
                statut_validation='pending'
            ).count(),
        }
    }
    return render(request, 'dashbord/rapports.html', context)


@login_required
def export_personnes_excel(request):
    """
    Exporte la liste des personnes en Excel
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Personnes"
    
    # En-têtes
    headers = [
        'Code', 'Nom', 'Prénom', 'Surnom', 'Genre', 'Date de naissance',
        'Lieu de naissance', 'Nationalité', 'CNI', 'Téléphone', 'Email',
        'Situation matrimoniale', 'Est vivant', 'Date de décès',
        'Type de résidence', 'Lieu de résidence', 'Famille', 'Village',
        'Âge', 'Date de création'
    ]
    
    # Style pour les en-têtes
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Appliquer les en-têtes
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Récupérer les données
    personnes = _visible_persons(request).select_related('famille', 'famille__village')
    
    # Remplir les données
    for row_num, personne in enumerate(personnes, 2):
        ws.cell(row=row_num, column=1, value=personne.code or '')
        ws.cell(row=row_num, column=2, value=personne.nom)
        ws.cell(row=row_num, column=3, value=personne.prenom)
        ws.cell(row=row_num, column=4, value=personne.surnom)
        ws.cell(row=row_num, column=5, value=personne.get_genre_display())
        ws.cell(row=row_num, column=6, value=personne.date_naissance.strftime('%d/%m/%Y') if personne.date_naissance else '')
        ws.cell(row=row_num, column=7, value=personne.lieu_naissance)
        ws.cell(row=row_num, column=8, value=personne.nationalite)
        ws.cell(row=row_num, column=9, value=personne.numero_cni)
        ws.cell(row=row_num, column=10, value=personne.telephone)
        ws.cell(row=row_num, column=11, value=personne.email)
        ws.cell(row=row_num, column=12, value=personne.get_situation_matrimoniale_display())
        ws.cell(row=row_num, column=13, value='Oui' if personne.est_vivant else 'Non')
        ws.cell(row=row_num, column=14, value=personne.date_deces.strftime('%d/%m/%Y') if personne.date_deces else '')
        ws.cell(row=row_num, column=15, value=personne.get_type_residence_display())
        ws.cell(row=row_num, column=16, value=personne.lieu_residence)
        ws.cell(row=row_num, column=17, value=personne.famille.nom_famille if personne.famille else '')
        ws.cell(row=row_num, column=18, value=personne.famille.village.nom if personne.famille and personne.famille.village else '')
        ws.cell(row=row_num, column=19, value=personne.age)
        ws.cell(row=row_num, column=20, value=personne.date_creation.strftime('%d/%m/%Y %H:%M'))
    
    # Ajuster la largeur des colonnes
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Créer la réponse HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=personnes_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(response)
    
    return response


@login_required
def export_familles_excel(request):
    """
    Exporte la liste des familles en Excel
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Familles"
    
    # En-têtes
    headers = [
        'Nom de famille', 'Village', 'Description', 'Nombre de membres',
        'Nombre d\'hommes', 'Nombre de femmes', 'Nombre de vivants',
        'Nombre en diaspora', 'Chef de famille', 'Date de création'
    ]
    
    # Style pour les en-têtes
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Appliquer les en-têtes
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Récupérer les données
    familles = _visible_families(request).select_related('village').prefetch_related('membres')
    
    # Remplir les données
    for row_num, famille in enumerate(familles, 2):
        if is_limited_data_entry_agent(request.user):
            membres_visibles = famille.membres.filter(
                deleted__isnull=True,
                created_by=request.user,
            )
            chef = membres_visibles.filter(est_chef_famille=True).first() or membres_visibles.first()
            nombre_membres = membres_visibles.count()
            nombre_hommes = membres_visibles.filter(genre='M').count()
            nombre_femmes = membres_visibles.filter(genre='F').count()
            nombre_vivants = membres_visibles.filter(est_vivant=True).count()
            nombre_diaspora = membres_visibles.filter(type_residence='diaspora').count()
        else:
            chef = famille.chef
            nombre_membres = famille.nombre_membres
            nombre_hommes = famille.nombre_hommes
            nombre_femmes = famille.nombre_femmes
            nombre_vivants = famille.nombre_vivants
            nombre_diaspora = famille.nombre_diaspora
        ws.cell(row=row_num, column=1, value=famille.nom_famille)
        ws.cell(row=row_num, column=2, value=famille.village.nom)
        ws.cell(row=row_num, column=3, value=famille.description)
        ws.cell(row=row_num, column=4, value=nombre_membres)
        ws.cell(row=row_num, column=5, value=nombre_hommes)
        ws.cell(row=row_num, column=6, value=nombre_femmes)
        ws.cell(row=row_num, column=7, value=nombre_vivants)
        ws.cell(row=row_num, column=8, value=nombre_diaspora)
        ws.cell(row=row_num, column=9, value=chef.nom_complet if chef else '')
        ws.cell(row=row_num, column=10, value=famille.date_creation.strftime('%d/%m/%Y %H:%M'))
    
    # Ajuster la largeur des colonnes
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20
    
    # Créer la réponse HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=familles_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(response)
    
    return response


@login_required
def export_villages_excel(request):
    """
    Exporte la liste des villages et leurs infrastructures en Excel
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Villages"
    
    # En-têtes
    headers = [
        'Nom du village', 'Description', 'Membres enregistrés', 'Chef du village',
        'Latitude', 'Longitude', 'Nombre de familles', 'Nombre d\'habitants',
        'Nombre d\'écoles', 'Nombre d\'hôpitaux', 'Nombre de dispensaires',
        'Nombre de centres de santé', 'Total infrastructures', 'Date de création'
    ]
    
    # Style pour les en-têtes
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Appliquer les en-têtes
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Récupérer les données
    villages = _visible_villages(request).prefetch_related('familles', 'infrastructures')
    
    # Remplir les données
    for row_num, village in enumerate(villages, 2):
        ws.cell(row=row_num, column=1, value=village.nom)
        ws.cell(row=row_num, column=2, value=village.description)
        ws.cell(row=row_num, column=3, value=village.nombre_habitants)
        ws.cell(row=row_num, column=4, value=village.chef_village)
        ws.cell(row=row_num, column=5, value=village.latitude)
        ws.cell(row=row_num, column=6, value=village.longitude)
        ws.cell(row=row_num, column=7, value=village.nombre_familles)
        ws.cell(row=row_num, column=8, value=village.nombre_habitants)
        ws.cell(row=row_num, column=9, value=village.nombre_ecoles)
        ws.cell(row=row_num, column=10, value=village.nombre_hopitaux)
        ws.cell(row=row_num, column=11, value=village.nombre_dispensaires)
        ws.cell(row=row_num, column=12, value=village.nombre_centres_sante)
        ws.cell(row=row_num, column=13, value=village.nombre_total_infrastructures)
        ws.cell(row=row_num, column=14, value=village.date_creation.strftime('%d/%m/%Y %H:%M'))
    
    # Ajuster la largeur des colonnes
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    
    # Créer la réponse HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=villages_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(response)
    
    return response


@login_required
def export_demographie_excel(request):
    """
    Exporte un rapport démographique complet en Excel
    """
    wb = openpyxl.Workbook()
    
    # Feuille 1: Statistiques générales
    ws1 = wb.active
    ws1.title = "Statistiques Générales"
    
    # En-têtes pour les statistiques
    stats_headers = ['Indicateur', 'Valeur', 'Pourcentage']
    
    # Style pour les en-têtes
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Appliquer les en-têtes
    for col_num, header in enumerate(stats_headers, 1):
        cell = ws1.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Calculer les statistiques
    visible_persons = _visible_persons(request)
    total_personnes = visible_persons.count()
    total_hommes = visible_persons.filter(genre='M').count()
    total_femmes = visible_persons.filter(genre='F').count()
    total_vivants = visible_persons.filter(est_vivant=True).count()
    total_decedes = visible_persons.filter(est_vivant=False).count()
    total_diaspora = visible_persons.filter(type_residence='diaspora').count()
    total_village = visible_persons.filter(type_residence='village').count()
    total_ci = visible_persons.filter(type_residence='ci').count()
    
    # Remplir les statistiques
    stats_data = [
        ('Total personnes', total_personnes, '100%'),
        ('Hommes', total_hommes, f'{(total_hommes/total_personnes*100):.1f}%' if total_personnes > 0 else '0%'),
        ('Femmes', total_femmes, f'{(total_femmes/total_personnes*100):.1f}%' if total_personnes > 0 else '0%'),
        ('Personnes vivantes', total_vivants, f'{(total_vivants/total_personnes*100):.1f}%' if total_personnes > 0 else '0%'),
        ('Personnes décédées', total_decedes, f'{(total_decedes/total_personnes*100):.1f}%' if total_personnes > 0 else '0%'),
        ('Diaspora', total_diaspora, f'{(total_diaspora/total_personnes*100):.1f}%' if total_personnes > 0 else '0%'),
        ('Résident au village', total_village, f'{(total_village/total_personnes*100):.1f}%' if total_personnes > 0 else '0%'),
        ('Résident CI (hors village)', total_ci, f'{(total_ci/total_personnes*100):.1f}%' if total_personnes > 0 else '0%'),
    ]
    
    for row_num, (indicateur, valeur, pourcentage) in enumerate(stats_data, 2):
        ws1.cell(row=row_num, column=1, value=indicateur)
        ws1.cell(row=row_num, column=2, value=valeur)
        ws1.cell(row=row_num, column=3, value=pourcentage)
    
    # Ajuster la largeur des colonnes
    ws1.column_dimensions['A'].width = 30
    ws1.column_dimensions['B'].width = 15
    ws1.column_dimensions['C'].width = 15
    
    # Feuille 2: Répartition par village
    ws2 = wb.create_sheet("Répartition par Village")
    
    village_headers = ['Village', 'Nombre de familles', 'Nombre d\'habitants', 'Hommes', 'Femmes', 'Diaspora']
    
    # Appliquer les en-têtes pour la feuille 2
    for col_num, header in enumerate(village_headers, 1):
        cell = ws2.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Récupérer les données par village
    villages = _visible_villages(request).prefetch_related('familles__membres')
    
    for row_num, village in enumerate(villages, 2):
        total_habitants_village = 0
        hommes_village = 0
        femmes_village = 0
        diaspora_village = 0
        
        for famille in village.familles.all():
            membres = famille.membres.filter(deleted__isnull=True)
            if is_limited_data_entry_agent(request.user):
                membres = membres.filter(created_by=request.user)
            total_habitants_village += membres.count()
            hommes_village += membres.filter(genre='M').count()
            femmes_village += membres.filter(genre='F').count()
            diaspora_village += membres.filter(type_residence='diaspora').count()
        
        ws2.cell(row=row_num, column=1, value=village.nom)
        ws2.cell(row=row_num, column=2, value=village.nombre_familles)
        ws2.cell(row=row_num, column=3, value=total_habitants_village)
        ws2.cell(row=row_num, column=4, value=hommes_village)
        ws2.cell(row=row_num, column=5, value=femmes_village)
        ws2.cell(row=row_num, column=6, value=diaspora_village)
    
    # Ajuster la largeur des colonnes pour la feuille 2
    for col in range(1, len(village_headers) + 1):
        ws2.column_dimensions[get_column_letter(col)].width = 20
    
    # Créer la réponse HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=rapport_demographie_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(response)
    
    return response


def _report_header_styles():
    return (
        Font(bold=True, color="FFFFFF"),
        PatternFill(start_color="366092", end_color="366092", fill_type="solid"),
        Alignment(horizontal="center", vertical="center"),
    )


def _apply_headers(ws, headers):
    header_font, header_fill, header_alignment = _report_header_styles()
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment


def _set_column_widths(ws, headers, default_width=18):
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = default_width


def _visible_cotisation_persons(request):
    return _visible_persons(request).filter(est_vivant=True).select_related(
        'famille',
        'famille__village',
    )


def _applicable_cotisations_for_person(person, cotisations):
    rows = []
    for cotisation in cotisations:
        if cotisation.est_generale:
            rows.append(cotisation)
        elif cotisation.famille_id == person.famille_id:
            rows.append(cotisation)
        elif (
            cotisation.famille_id is None
            and cotisation.village_id == person.famille.village_id
        ):
            rows.append(cotisation)
    return rows


def _group_cotisation_payments(persons, cotisations):
    grouped = {}
    for payment in (
        Paiement.objects.select_related(
            'personne',
            'cotisation',
            'compte_paiement',
        )
        .filter(personne__in=persons, cotisation__in=cotisations)
        .order_by('-date_paiement', '-date_creation')
    ):
        grouped.setdefault((str(payment.personne_id), str(payment.cotisation_id)), []).append(payment)
    return grouped


def _group_cotisation_trackings(persons, cotisations):
    grouped = {}
    for tracking in CotisationPersonne.objects.filter(
        personne__in=persons,
        cotisation__in=cotisations,
    ).select_related('personne', 'cotisation'):
        grouped[(str(tracking.personne_id), str(tracking.cotisation_id))] = tracking
    return grouped


def _payment_breakdown(payment_list):
    approved = [item for item in payment_list if item.statut_validation == 'approved']
    pending = [item for item in payment_list if item.statut_validation == 'pending']
    rejected = [item for item in payment_list if item.statut_validation == 'rejected']
    total_paid = sum((item.montant for item in approved), Decimal('0.00'))
    return approved, pending, rejected, total_paid


@login_required
def export_cotisations_personnes_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cotisations par personne"

    selected_cotisation_id = request.GET.get('cotisation', '').strip()
    headers = [
        'Code personne', 'Nom complet', 'Famille', 'Village', 'Telephone', 'Email',
        'Periode', 'Cible', 'Montant attendu', 'Montant valide', 'Reste a payer',
        'Statut suivi', 'Nb paiements', 'Paiements en attente', 'Dernier paiement',
        'Dernier mode', 'Compte de depot', 'Derniere reference',
    ]
    _apply_headers(ws, headers)

    persons = list(_visible_cotisation_persons(request))
    cotisations_qs = Cotisation.objects.select_related(
        'village',
        'famille',
        'famille__village',
    ).order_by('-annee', '-mois')
    if selected_cotisation_id:
        cotisations_qs = cotisations_qs.filter(id=selected_cotisation_id)
    cotisations = list(cotisations_qs)
    payments = _group_cotisation_payments(persons, cotisations)
    trackings = _group_cotisation_trackings(persons, cotisations)

    row_num = 2
    for person in persons:
        for cotisation in _applicable_cotisations_for_person(person, cotisations):
            payment_list = payments.get((str(person.id), str(cotisation.id)), [])
            tracking = trackings.get((str(person.id), str(cotisation.id)))
            approved, pending, rejected, total_paid = _payment_breakdown(payment_list)
            expected_amount = (
                tracking.montant_attendu
                if tracking and tracking.montant_attendu is not None
                else None
            )
            remaining_amount = compute_remaining_amount(expected_amount, total_paid)
            status_key = resolve_person_tracking_status(
                expected_amount,
                total_paid,
                pending_count=len(pending),
                rejected_count=len(rejected),
            )
            follow_up_status = PERSON_TRACKING_STATUS_LABELS[status_key]
            last_payment = payment_list[0] if payment_list else None

            ws.cell(row=row_num, column=1, value=person.code or '')
            ws.cell(row=row_num, column=2, value=person.nom_complet)
            ws.cell(row=row_num, column=3, value=person.famille.nom_famille if person.famille else '')
            ws.cell(row=row_num, column=4, value=person.famille.village.nom if person.famille and person.famille.village else '')
            ws.cell(row=row_num, column=5, value=person.telephone)
            ws.cell(row=row_num, column=6, value=person.email)
            ws.cell(row=row_num, column=7, value=cotisation.periode_label)
            ws.cell(row=row_num, column=8, value=cotisation.cible_label)
            ws.cell(
                row=row_num,
                column=9,
                value=float(expected_amount) if expected_amount is not None else '',
            )
            ws.cell(row=row_num, column=10, value=float(total_paid))
            ws.cell(
                row=row_num,
                column=11,
                value=float(remaining_amount) if remaining_amount is not None else '',
            )
            ws.cell(row=row_num, column=12, value=follow_up_status)
            ws.cell(row=row_num, column=13, value=len(payment_list))
            ws.cell(row=row_num, column=14, value=len(pending))
            ws.cell(
                row=row_num,
                column=15,
                value=last_payment.date_paiement.strftime('%d/%m/%Y') if last_payment else '',
            )
            ws.cell(
                row=row_num,
                column=16,
                value=last_payment.get_mode_paiement_display() if last_payment else '',
            )
            ws.cell(
                row=row_num,
                column=17,
                value=last_payment.compte_paiement.numero if last_payment and last_payment.compte_paiement else '',
            )
            ws.cell(
                row=row_num,
                column=18,
                value=last_payment.reference_transaction if last_payment else '',
            )
            row_num += 1

    _set_column_widths(ws, headers, default_width=20)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = (
        f'attachment; filename=cotisations_personnes_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )
    wb.save(response)
    return response


@login_required
def export_cotisations_villages_excel(request):
    wb = openpyxl.Workbook()
    ws_villages = wb.active
    ws_villages.title = "Par village"
    village_headers = [
        'Village', 'Cotisations', 'Personnes cibles', 'Payeurs uniques',
        'Total collecte', 'Paiements en attente',
    ]
    _apply_headers(ws_villages, village_headers)

    visible_persons = _visible_cotisation_persons(request)
    villages = _visible_villages(request).order_by('nom')
    for row_num, village in enumerate(villages, 2):
        cotisations = Cotisation.objects.filter(village=village)
        payeurs = Paiement.objects.filter(
            cotisation__village=village,
            statut_validation='approved',
            personne__in=visible_persons,
        ).values('personne_id').distinct().count()
        total_collecte = (
            Paiement.objects.filter(
                cotisation__village=village,
                statut_validation='approved',
                personne__in=visible_persons,
            ).aggregate(total=Sum('montant'))['total']
            or 0
        )
        pending_total = Paiement.objects.filter(
            cotisation__village=village,
            statut_validation='pending',
            personne__in=visible_persons,
        ).count()

        ws_villages.cell(row=row_num, column=1, value=village.nom)
        ws_villages.cell(row=row_num, column=2, value=cotisations.count())
        ws_villages.cell(
            row=row_num,
            column=3,
            value=visible_persons.filter(famille__village=village).count(),
        )
        ws_villages.cell(row=row_num, column=4, value=payeurs)
        ws_villages.cell(row=row_num, column=5, value=float(total_collecte))
        ws_villages.cell(row=row_num, column=6, value=pending_total)

    ws_familles = wb.create_sheet("Par famille")
    family_headers = [
        'Famille', 'Village', 'Cotisations', 'Membres vivants',
        'Payeurs uniques', 'Total collecte', 'Paiements en attente',
    ]
    _apply_headers(ws_familles, family_headers)

    familles = _visible_families(request).select_related('village').order_by('nom_famille')
    for row_num, famille in enumerate(familles, 2):
        cotisations = Cotisation.objects.filter(famille=famille)
        payeurs = Paiement.objects.filter(
            cotisation__famille=famille,
            statut_validation='approved',
            personne__in=visible_persons,
        ).values('personne_id').distinct().count()
        total_collecte = (
            Paiement.objects.filter(
                cotisation__famille=famille,
                statut_validation='approved',
                personne__in=visible_persons,
            ).aggregate(total=Sum('montant'))['total']
            or 0
        )
        pending_total = Paiement.objects.filter(
            cotisation__famille=famille,
            statut_validation='pending',
            personne__in=visible_persons,
        ).count()

        ws_familles.cell(row=row_num, column=1, value=famille.nom_famille)
        ws_familles.cell(row=row_num, column=2, value=famille.village.nom if famille.village else '')
        ws_familles.cell(row=row_num, column=3, value=cotisations.count())
        ws_familles.cell(
            row=row_num,
            column=4,
            value=visible_persons.filter(famille=famille).count(),
        )
        ws_familles.cell(row=row_num, column=5, value=payeurs)
        ws_familles.cell(row=row_num, column=6, value=float(total_collecte))
        ws_familles.cell(row=row_num, column=7, value=pending_total)

    _set_column_widths(ws_villages, village_headers, default_width=20)
    _set_column_widths(ws_familles, family_headers, default_width=20)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = (
        f'attachment; filename=cotisations_villages_familles_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )
    wb.save(response)
    return response


@login_required
def export_cotisations_annuelles_excel(request):
    year = request.GET.get('annee', '').strip()
    try:
        selected_year = int(year) if year else timezone.now().year
    except ValueError:
        selected_year = timezone.now().year

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Annuel {selected_year}"
    headers = [
        'Mois', 'Cotisations ouvertes', 'Cotisations fermees', 'Payeurs uniques',
        'Total collecte', 'Wave', 'Orange Money', 'Moov', 'MTN', 'Virement', 'Espece',
    ]
    _apply_headers(ws, headers)

    visible_persons = _visible_cotisation_persons(request)
    for month in range(1, 13):
        monthly_cotisations = Cotisation.objects.filter(annee=selected_year, mois=month)
        monthly_payments = Paiement.objects.filter(
            cotisation__annee=selected_year,
            cotisation__mois=month,
            statut_validation='approved',
            personne__in=visible_persons,
        )
        row = month + 1
        ws.cell(row=row, column=1, value=date(2000, month, 1).strftime('%B'))
        ws.cell(row=row, column=2, value=monthly_cotisations.filter(statut='ouverte').count())
        ws.cell(row=row, column=3, value=monthly_cotisations.filter(statut='fermee').count())
        ws.cell(row=row, column=4, value=monthly_payments.values('personne_id').distinct().count())
        ws.cell(row=row, column=5, value=float(monthly_payments.aggregate(total=Sum('montant'))['total'] or 0))
        ws.cell(row=row, column=6, value=float(monthly_payments.filter(mode_paiement='wave').aggregate(total=Sum('montant'))['total'] or 0))
        ws.cell(row=row, column=7, value=float(monthly_payments.filter(mode_paiement='orange_money').aggregate(total=Sum('montant'))['total'] or 0))
        ws.cell(row=row, column=8, value=float(monthly_payments.filter(mode_paiement='moov').aggregate(total=Sum('montant'))['total'] or 0))
        ws.cell(row=row, column=9, value=float(monthly_payments.filter(mode_paiement='mtn').aggregate(total=Sum('montant'))['total'] or 0))
        ws.cell(row=row, column=10, value=float(monthly_payments.filter(mode_paiement='virement').aggregate(total=Sum('montant'))['total'] or 0))
        ws.cell(row=row, column=11, value=float(monthly_payments.filter(mode_paiement='espece').aggregate(total=Sum('montant'))['total'] or 0))

    ws_detail = wb.create_sheet("Campagnes")
    detail_headers = [
        'Periode', 'Cible', 'Statut', 'Payeurs', 'Total collecte',
        'Montant attendu', 'Reste estime', 'Sans paiement',
    ]
    _apply_headers(ws_detail, detail_headers)
    annual_cotisations = Cotisation.objects.filter(annee=selected_year).select_related(
        'village',
        'famille',
        'famille__village',
    ).order_by('mois')
    for row_num, cotisation in enumerate(annual_cotisations, 2):
        ws_detail.cell(row=row_num, column=1, value=cotisation.periode_label)
        ws_detail.cell(row=row_num, column=2, value=cotisation.cible_label)
        ws_detail.cell(row=row_num, column=3, value=cotisation.get_statut_display())
        ws_detail.cell(row=row_num, column=4, value=cotisation.nombre_payeurs)
        ws_detail.cell(row=row_num, column=5, value=float(cotisation.total_collecte))
        ws_detail.cell(row=row_num, column=6, value=float(cotisation.total_attendu))
        ws_detail.cell(row=row_num, column=7, value=float(cotisation.reste_global))
        ws_detail.cell(row=row_num, column=8, value=cotisation.personnes_sans_paiement.count())

    _set_column_widths(ws, headers, default_width=18)
    _set_column_widths(ws_detail, detail_headers, default_width=22)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = (
        f'attachment; filename=cotisations_annuelles_{selected_year}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )
    wb.save(response)
    return response


@login_required
def export_relances_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Relances"
    headers = [
        'Periode', 'Cible', 'Nom complet', 'Famille', 'Village',
        'Telephone', 'Email', 'Montant attendu', 'Montant valide',
        'Reste a payer', 'Statut suivi', 'Type de residence',
    ]
    _apply_headers(ws, headers)

    visible_persons = list(_visible_cotisation_persons(request))
    row_num = 2
    cotisations = list(Cotisation.objects.filter(statut='ouverte').select_related(
        'village',
        'famille',
        'famille__village',
    ).order_by('-annee', '-mois'))
    payments = _group_cotisation_payments(visible_persons, cotisations)
    trackings = _group_cotisation_trackings(visible_persons, cotisations)

    for person in visible_persons:
        for cotisation in _applicable_cotisations_for_person(person, cotisations):
            payment_list = payments.get((str(person.id), str(cotisation.id)), [])
            tracking = trackings.get((str(person.id), str(cotisation.id)))
            approved, pending, rejected, total_paid = _payment_breakdown(payment_list)
            expected_amount = (
                tracking.montant_attendu
                if tracking and tracking.montant_attendu is not None
                else None
            )
            remaining_amount = compute_remaining_amount(expected_amount, total_paid)
            status_key = resolve_person_tracking_status(
                expected_amount,
                total_paid,
                pending_count=len(pending),
                rejected_count=len(rejected),
            )
            if status_key in {'solde', 'versement'}:
                continue

            ws.cell(row=row_num, column=1, value=cotisation.periode_label)
            ws.cell(row=row_num, column=2, value=cotisation.cible_label)
            ws.cell(row=row_num, column=3, value=person.nom_complet)
            ws.cell(row=row_num, column=4, value=person.famille.nom_famille if person.famille else '')
            ws.cell(row=row_num, column=5, value=person.famille.village.nom if person.famille and person.famille.village else '')
            ws.cell(row=row_num, column=6, value=person.telephone)
            ws.cell(row=row_num, column=7, value=person.email)
            ws.cell(
                row=row_num,
                column=8,
                value=float(expected_amount) if expected_amount is not None else '',
            )
            ws.cell(row=row_num, column=9, value=float(total_paid))
            ws.cell(
                row=row_num,
                column=10,
                value=float(remaining_amount) if remaining_amount is not None else '',
            )
            ws.cell(row=row_num, column=11, value=PERSON_TRACKING_STATUS_LABELS[status_key])
            ws.cell(row=row_num, column=12, value=person.get_type_residence_display())
            row_num += 1

    _set_column_widths(ws, headers, default_width=22)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = (
        f'attachment; filename=relances_cotisations_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )
    wb.save(response)
    return response
