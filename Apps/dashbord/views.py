from django.shortcuts import render
from django.http import HttpResponse
from django.db.models import Count, Q, Avg, Sum
from django.utils import timezone
from datetime import datetime, date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from Apps.person.models import Person
from Apps.families.models import Family
from Apps.villages.models import Village, Infrastructure

def rapports_dashboard(request):
    """
    Vue principale pour la page des rapports
    """
    context = {
        'title': 'Rapports et Exportations',
        'stats': {
            'total_personnes': Person.objects.filter(deleted__isnull=True).count(),
            'total_familles': Family.objects.filter(deleted__isnull=True).count(),
            'total_villages': Village.objects.filter(deleted__isnull=True).count(),
            'total_infrastructures': Infrastructure.objects.filter(deleted__isnull=True).count(),
        }
    }
    return render(request, 'dashbord/rapports.html', context)

def export_personnes_excel(request):
    """
    Exporte la liste des personnes en Excel
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Personnes"
    
    # En-têtes
    headers = [
        'Code', 'Nom', 'Prénom', 'Surnom', 'Genre', 'Date de naissance',
        'Lieu de naissance', 'Nationalité', 'CNI', 'Téléphone', 'Email',
        'Situation matrimoniale', 'Est vivant', 'Date de décès',
        'Type de résidence', 'Lieu de résidence', 'Famille', 'Village',
        'Âge', 'Date de création'
    ]
    
    # Style pour les en-têtes
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Appliquer les en-têtes
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Récupérer les données
    personnes = Person.objects.filter(deleted__isnull=True).select_related('famille', 'famille__village')
    
    # Remplir les données
    for row_num, personne in enumerate(personnes, 2):
        ws.cell(row=row_num, column=1, value=personne.code or '')
        ws.cell(row=row_num, column=2, value=personne.nom)
        ws.cell(row=row_num, column=3, value=personne.prenom)
        ws.cell(row=row_num, column=4, value=personne.surnom)
        ws.cell(row=row_num, column=5, value=personne.get_genre_display())
        ws.cell(row=row_num, column=6, value=personne.date_naissance.strftime('%d/%m/%Y') if personne.date_naissance else '')
        ws.cell(row=row_num, column=7, value=personne.lieu_naissance)
        ws.cell(row=row_num, column=8, value=personne.nationalite)
        ws.cell(row=row_num, column=9, value=personne.numero_cni)
        ws.cell(row=row_num, column=10, value=personne.telephone)
        ws.cell(row=row_num, column=11, value=personne.email)
        ws.cell(row=row_num, column=12, value=personne.get_situation_matrimoniale_display())
        ws.cell(row=row_num, column=13, value='Oui' if personne.est_vivant else 'Non')
        ws.cell(row=row_num, column=14, value=personne.date_deces.strftime('%d/%m/%Y') if personne.date_deces else '')
        ws.cell(row=row_num, column=15, value=personne.get_type_residence_display())
        ws.cell(row=row_num, column=16, value=personne.lieu_residence)
        ws.cell(row=row_num, column=17, value=personne.famille.nom_famille if personne.famille else '')
        ws.cell(row=row_num, column=18, value=personne.famille.village.nom if personne.famille and personne.famille.village else '')
        ws.cell(row=row_num, column=19, value=personne.age)
        ws.cell(row=row_num, column=20, value=personne.date_creation.strftime('%d/%m/%Y %H:%M'))
    
    # Ajuster la largeur des colonnes
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Créer la réponse HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=personnes_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(response)
    
    return response

def export_familles_excel(request):
    """
    Exporte la liste des familles en Excel
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Familles"
    
    # En-têtes
    headers = [
        'Nom de famille', 'Village', 'Description', 'Nombre de membres',
        'Nombre d\'hommes', 'Nombre de femmes', 'Nombre de vivants',
        'Nombre en diaspora', 'Chef de famille', 'Date de création'
    ]
    
    # Style pour les en-têtes
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Appliquer les en-têtes
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Récupérer les données
    familles = Family.objects.filter(deleted__isnull=True).select_related('village').prefetch_related('membres')
    
    # Remplir les données
    for row_num, famille in enumerate(familles, 2):
        chef = famille.chef
        ws.cell(row=row_num, column=1, value=famille.nom_famille)
        ws.cell(row=row_num, column=2, value=famille.village.nom)
        ws.cell(row=row_num, column=3, value=famille.description)
        ws.cell(row=row_num, column=4, value=famille.nombre_membres)
        ws.cell(row=row_num, column=5, value=famille.nombre_hommes)
        ws.cell(row=row_num, column=6, value=famille.nombre_femmes)
        ws.cell(row=row_num, column=7, value=famille.nombre_vivants)
        ws.cell(row=row_num, column=8, value=famille.nombre_diaspora)
        ws.cell(row=row_num, column=9, value=chef.nom_complet if chef else '')
        ws.cell(row=row_num, column=10, value=famille.date_creation.strftime('%d/%m/%Y %H:%M'))
    
    # Ajuster la largeur des colonnes
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20
    
    # Créer la réponse HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=familles_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(response)
    
    return response

def export_villages_excel(request):
    """
    Exporte la liste des villages et leurs infrastructures en Excel
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Villages"
    
    # En-têtes
    headers = [
        'Nom du village', 'Description', 'Population estimée', 'Chef du village',
        'Latitude', 'Longitude', 'Nombre de familles', 'Nombre d\'habitants',
        'Nombre d\'écoles', 'Nombre d\'hôpitaux', 'Nombre de dispensaires',
        'Nombre de centres de santé', 'Total infrastructures', 'Date de création'
    ]
    
    # Style pour les en-têtes
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Appliquer les en-têtes
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Récupérer les données
    villages = Village.objects.filter(deleted__isnull=True).prefetch_related('familles', 'infrastructures')
    
    # Remplir les données
    for row_num, village in enumerate(villages, 2):
        ws.cell(row=row_num, column=1, value=village.nom)
        ws.cell(row=row_num, column=2, value=village.description)
        ws.cell(row=row_num, column=3, value=village.population_estimee)
        ws.cell(row=row_num, column=4, value=village.chef_village)
        ws.cell(row=row_num, column=5, value=village.latitude)
        ws.cell(row=row_num, column=6, value=village.longitude)
        ws.cell(row=row_num, column=7, value=village.nombre_familles)
        ws.cell(row=row_num, column=8, value=village.nombre_habitants)
        ws.cell(row=row_num, column=9, value=village.nombre_ecoles)
        ws.cell(row=row_num, column=10, value=village.nombre_hopitaux)
        ws.cell(row=row_num, column=11, value=village.nombre_dispensaires)
        ws.cell(row=row_num, column=12, value=village.nombre_centres_sante)
        ws.cell(row=row_num, column=13, value=village.nombre_total_infrastructures)
        ws.cell(row=row_num, column=14, value=village.date_creation.strftime('%d/%m/%Y %H:%M'))
    
    # Ajuster la largeur des colonnes
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    
    # Créer la réponse HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=villages_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(response)
    
    return response

def export_demographie_excel(request):
    """
    Exporte un rapport démographique complet en Excel
    """
    wb = openpyxl.Workbook()
    
    # Feuille 1: Statistiques générales
    ws1 = wb.active
    ws1.title = "Statistiques Générales"
    
    # En-têtes pour les statistiques
    stats_headers = ['Indicateur', 'Valeur', 'Pourcentage']
    
    # Style pour les en-têtes
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Appliquer les en-têtes
    for col_num, header in enumerate(stats_headers, 1):
        cell = ws1.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Calculer les statistiques
    total_personnes = Person.objects.filter(deleted__isnull=True).count()
    total_hommes = Person.objects.filter(deleted__isnull=True, genre='M').count()
    total_femmes = Person.objects.filter(deleted__isnull=True, genre='F').count()
    total_vivants = Person.objects.filter(deleted__isnull=True, est_vivant=True).count()
    total_decedes = Person.objects.filter(deleted__isnull=True, est_vivant=False).count()
    total_diaspora = Person.objects.filter(deleted__isnull=True, type_residence='diaspora').count()
    total_village = Person.objects.filter(deleted__isnull=True, type_residence='village').count()
    total_ci = Person.objects.filter(deleted__isnull=True, type_residence='ci').count()
    
    # Remplir les statistiques
    stats_data = [
        ('Total personnes', total_personnes, '100%'),
        ('Hommes', total_hommes, f'{(total_hommes/total_personnes*100):.1f}%' if total_personnes > 0 else '0%'),
        ('Femmes', total_femmes, f'{(total_femmes/total_personnes*100):.1f}%' if total_personnes > 0 else '0%'),
        ('Personnes vivantes', total_vivants, f'{(total_vivants/total_personnes*100):.1f}%' if total_personnes > 0 else '0%'),
        ('Personnes décédées', total_decedes, f'{(total_decedes/total_personnes*100):.1f}%' if total_personnes > 0 else '0%'),
        ('Diaspora', total_diaspora, f'{(total_diaspora/total_personnes*100):.1f}%' if total_personnes > 0 else '0%'),
        ('Résident au village', total_village, f'{(total_village/total_personnes*100):.1f}%' if total_personnes > 0 else '0%'),
        ('Résident CI (hors village)', total_ci, f'{(total_ci/total_personnes*100):.1f}%' if total_personnes > 0 else '0%'),
    ]
    
    for row_num, (indicateur, valeur, pourcentage) in enumerate(stats_data, 2):
        ws1.cell(row=row_num, column=1, value=indicateur)
        ws1.cell(row=row_num, column=2, value=valeur)
        ws1.cell(row=row_num, column=3, value=pourcentage)
    
    # Ajuster la largeur des colonnes
    ws1.column_dimensions['A'].width = 30
    ws1.column_dimensions['B'].width = 15
    ws1.column_dimensions['C'].width = 15
    
    # Feuille 2: Répartition par village
    ws2 = wb.create_sheet("Répartition par Village")
    
    village_headers = ['Village', 'Nombre de familles', 'Nombre d\'habitants', 'Hommes', 'Femmes', 'Diaspora']
    
    # Appliquer les en-têtes pour la feuille 2
    for col_num, header in enumerate(village_headers, 1):
        cell = ws2.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Récupérer les données par village
    villages = Village.objects.filter(deleted__isnull=True).prefetch_related('familles__membres')
    
    for row_num, village in enumerate(villages, 2):
        total_habitants_village = 0
        hommes_village = 0
        femmes_village = 0
        diaspora_village = 0
        
        for famille in village.familles.all():
            membres = famille.membres.filter(deleted__isnull=True)
            total_habitants_village += membres.count()
            hommes_village += membres.filter(genre='M').count()
            femmes_village += membres.filter(genre='F').count()
            diaspora_village += membres.filter(type_residence='diaspora').count()
        
        ws2.cell(row=row_num, column=1, value=village.nom)
        ws2.cell(row=row_num, column=2, value=village.nombre_familles)
        ws2.cell(row=row_num, column=3, value=total_habitants_village)
        ws2.cell(row=row_num, column=4, value=hommes_village)
        ws2.cell(row=row_num, column=5, value=femmes_village)
        ws2.cell(row=row_num, column=6, value=diaspora_village)
    
    # Ajuster la largeur des colonnes pour la feuille 2
    for col in range(1, len(village_headers) + 1):
        ws2.column_dimensions[get_column_letter(col)].width = 20
    
    # Créer la réponse HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=rapport_demographie_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(response)
    
    return response
